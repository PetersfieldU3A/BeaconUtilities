"""Sync orchestration for BeaconUtilities.

Coordinates the full Phase I Beacon → WordPress workflow:

1. Preflight validation
2. Beacon login + Excel export download (via Playwright)
3. Excel parsing → :class:`~beaconutilities.models.BeaconRecord` lists
4. Optional SQLite staging — all workbook sheets loaded as tables
5. Field mapping → WordPress REST API payloads
6. WordPress upsert (skipped in dry-run mode)
7. State persistence
"""

from __future__ import annotations

from datetime import datetime
import logging
from pathlib import Path

import openpyxl

from .beacon_scraper import download_beacon_backup, download_beacon_exports
from .database import load_workbook_to_db
from .excel_parser import parse_sheet
from .mapping import map_record
from .models import BeaconRecord, EntityType
from .preflight import preflight_beacon, preflight_beacon_backup, preflight_wordpress
from .state import load_state, save_state
from .wordpress import client_from_config

log = logging.getLogger(__name__)

#: Default path for the runtime state file.
DEFAULT_STATE_PATH = Path("state/state.json")

_MEMBER_NAME_COLUMNS = ["mem_no", "status", "title", "forename", "surname"]


def run_sync(config: dict, dry_run: bool = False) -> dict:
    """Orchestrate a full Beacon → WordPress sync cycle.

    Args:
        config: Loaded configuration dictionary produced by
                :func:`~beaconutilities.config.load_config`.
        dry_run: When ``True`` data is extracted and mapped but nothing is
                 written to WordPress.  State is not updated in dry-run mode.

    Returns:
        Result summary dict with keys:
        ``status``, ``dry_run``, ``members_extracted``, ``groups_extracted``,
        ``published``, ``errors``.
    """
    log.info("Sync started (dry_run=%s)", dry_run)

    result: dict = {
        "status": "ok",
        "dry_run": dry_run,
        "members_extracted": 0,
        "groups_extracted": 0,
        "staged": 0,
        "published": 0,
        "errors": [],
    }

    # ── 1. Preflight ─────────────────────────────────────────────────────────
    if not preflight_beacon(config):
        result["status"] = "preflight_failed"
        return result
    if not dry_run and not preflight_wordpress(config):
        result["status"] = "preflight_failed"
        return result

    # ── 2. Download Excel exports ─────────────────────────────────────────────
    download_dir = Path(
        config.get("beacon_export", {}).get("download_dir", "downloads")
    )
    try:
        export_paths = download_beacon_exports(config, download_dir)
    except Exception as exc:
        log.error("Beacon export failed: %s", exc)
        result["status"] = "download_failed"
        result["errors"].append(str(exc))
        return result

    # ── 3. Parse ──────────────────────────────────────────────────────────────
    member_records = _parse_export(export_paths["members"], EntityType.MEMBER, result)
    group_records = _parse_export(export_paths["groups"], EntityType.GROUP, result)

    result["members_extracted"] = len(member_records)
    result["groups_extracted"] = len(group_records)
    log.info(
        "Extracted %d member(s) and %d group(s)",
        result["members_extracted"],
        result["groups_extracted"],
    )

    # ── 4. Optional staging to SQLite ───────────────────────────────────────
    db_cfg = config.get("database", {})
    db_enabled = _is_truthy(db_cfg.get("enabled", "false"))
    if db_enabled:
        db_path = Path(db_cfg.get("path", "state/beacon_data.db"))
        persist_across_sessions = _is_truthy(
            db_cfg.get("persist_across_sessions", "false")
        )
        try:
            counts = _stage_exports_to_db(
                export_paths,
                db_path,
                persist_across_sessions=persist_across_sessions,
            )
            result["staged"] = sum(counts.values())
            log.info(
                "Staged %d record(s) into %s (%d table(s), persist_across_sessions=%s)",
                result["staged"],
                db_path,
                len(counts),
                persist_across_sessions,
            )
        except Exception as exc:
            log.error("Database staging failed: %s", exc)
            result["errors"].append(f"db_staging: {exc}")
            result["status"] = "partial"

    if dry_run:
        log.info("Dry-run: skipping WordPress publish and state update")
        return result

    # ── 5 & 6. Map + Publish ──────────────────────────────────────────────────
    wp_client = client_from_config(config)
    wp_cfg = config.get("wordpress", {})
    members_post_type = wp_cfg.get("members_post_type", "posts")
    groups_post_type = wp_cfg.get("groups_post_type", "posts")

    all_records = [
        (record, members_post_type) for record in member_records
    ] + [
        (record, groups_post_type) for record in group_records
    ]

    for record, post_type in all_records:
        try:
            payload = map_record(record)
            wp_client.upsert_post(payload, post_type=post_type)
            result["published"] += 1
        except Exception as exc:
            msg = f"{record.entity_type.value} id={record.record_id}: {exc}"
            log.error("Publish failed — %s", msg)
            result["errors"].append(msg)
            result["status"] = "partial"

    # ── 7. State ──────────────────────────────────────────────────────────────
    state = load_state(DEFAULT_STATE_PATH)
    state["last_sync"] = {
        "status": result["status"],
        "members_extracted": result["members_extracted"],
        "groups_extracted": result["groups_extracted"],
        "published": result["published"],
        "errors": result["errors"],
    }
    save_state(DEFAULT_STATE_PATH, state)

    log.info("Sync complete: %s", result)
    return result


def run_beacon_to_sqlite_dry_run(config: dict) -> dict:
    """Download Beacon exports and stage all workbook sheets into SQLite.

    This task does not parse/map/publish data to WordPress and does not update
    state.json. It is intended as a fast validation and local staging step.
    """
    result: dict = {
        "status": "ok",
        "members_extracted": 0,
        "groups_extracted": 0,
        "staged": 0,
        "tables": 0,
        "errors": [],
    }

    if not preflight_beacon(config):
        result["status"] = "preflight_failed"
        return result

    download_dir = Path(config.get("beacon_export", {}).get("download_dir", "downloads"))
    try:
        export_paths = download_beacon_exports(config, download_dir)
    except Exception as exc:
        log.error("Beacon export failed: %s", exc)
        result["status"] = "download_failed"
        result["errors"].append(str(exc))
        return result

    result["members_extracted"] = _count_rows_in_export(export_paths["members"])
    result["groups_extracted"] = _count_rows_in_export(export_paths["groups"])

    db_cfg = config.get("database", {})
    db_path = Path(db_cfg.get("path", "state/beacon_data.db"))
    persist_across_sessions = _is_truthy(db_cfg.get("persist_across_sessions", "false"))
    try:
        counts = _stage_exports_to_db(
            export_paths,
            db_path,
            persist_across_sessions=persist_across_sessions,
        )
        result["staged"] = sum(counts.values())
        result["tables"] = len(counts)
    except Exception as exc:
        log.error("Database staging failed: %s", exc)
        result["status"] = "partial"
        result["errors"].append(f"db_staging: {exc}")

    return result


def run_export_member_names(config: dict, output_dir: Path) -> dict:
    """Download Beacon exports and write Member_Names.xlsx to output_dir."""
    result: dict = {
        "status": "ok",
        "rows_written": 0,
        "output_file": str(output_dir / "Member_Names.xlsx"),
        "errors": [],
    }

    if not preflight_beacon(config):
        result["status"] = "preflight_failed"
        return result

    download_dir = Path(config.get("beacon_export", {}).get("download_dir", "downloads"))
    try:
        export_paths = download_beacon_exports(config, download_dir)
    except Exception as exc:
        log.error("Beacon export failed: %s", exc)
        result["status"] = "download_failed"
        result["errors"].append(str(exc))
        return result

    members_file = export_paths["members"]
    try:
        rows = parse_sheet(members_file, "Members")
    except KeyError:
        rows = parse_sheet(members_file, 0)
    except Exception as exc:
        result["status"] = "parse_failed"
        result["errors"].append(str(exc))
        return result

    missing = [c for c in _MEMBER_NAME_COLUMNS if rows and c not in rows[0]]
    if missing:
        result["status"] = "parse_failed"
        result["errors"].append(f"Missing required member columns: {missing}")
        return result

    output_dir.mkdir(parents=True, exist_ok=True)
    out_file = output_dir / "Member_Names.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Member Names"
    ws.append(_MEMBER_NAME_COLUMNS)
    for row in rows:
        ws.append([row.get(c) for c in _MEMBER_NAME_COLUMNS])
    wb.save(out_file)
    wb.close()

    result["rows_written"] = len(rows)
    result["output_file"] = str(out_file)
    return result


def run_beacon_full_backup(config: dict, output_file: Path | None = None) -> dict:
    """Download Beacon full-backup workbook and save it locally.

    This task only downloads a single workbook and does not stage to SQLite,
    publish to WordPress, or update runtime state.
    """
    result: dict = {
        "status": "ok",
        "output_file": "",
        "errors": [],
    }

    if not preflight_beacon_backup(config):
        result["status"] = "preflight_failed"
        return result

    if output_file is None:
        backup_dir = Path(
            config.get("beacon_export", {}).get("backup_output_dir", "outputs")
        )
        destination = backup_dir / _default_backup_filename(config)
    else:
        destination = Path(output_file)
        if destination.suffix.lower() != ".xlsx":
            destination = destination / _default_backup_filename(config)

    try:
        saved_file = download_beacon_backup(config, destination)
    except Exception as exc:
        log.error("Beacon full-backup export failed: %s", exc)
        result["status"] = "download_failed"
        result["errors"].append(str(exc))
        return result

    result["output_file"] = str(saved_file)
    return result


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _parse_export(
    file_path: Path,
    entity_type: EntityType,
    result: dict,
) -> list[BeaconRecord]:
    """Parse all sheets of an Excel export file into BeaconRecord instances.

    Each sheet in the file contributes rows; the first column of each row is
    used as the ``record_id``.  Parsing errors are recorded in *result* but
    do not abort the sync.
    """
    from .excel_parser import list_sheets

    records: list[BeaconRecord] = []
    try:
        sheets = list_sheets(file_path)
    except Exception as exc:
        log.error("Cannot read %s: %s", file_path, exc)
        result["errors"].append(str(exc))
        return records

    for sheet in sheets:
        try:
            rows = parse_sheet(file_path, sheet)
        except Exception as exc:
            log.warning("Skipping sheet '%s' in %s: %s", sheet, file_path, exc)
            result["errors"].append(f"Sheet '{sheet}': {exc}")
            continue

        for row in rows:
            # Use the first non-empty column value as the record ID
            record_id = str(next((v for v in row.values() if v is not None), ""))
            records.append(
                BeaconRecord(
                    record_id=record_id,
                    entity_type=entity_type,
                    fields=row,
                )
            )

    return records


def _stage_exports_to_db(
    export_paths: dict[str, Path],
    db_path: Path,
    *,
    persist_across_sessions: bool,
) -> dict[str, int]:
    """Load each downloaded workbook into SQLite and return table row counts."""
    counts: dict[str, int] = {}
    for xlsx_path in export_paths.values():
        counts.update(
            load_workbook_to_db(
                db_path,
                xlsx_path,
                append=persist_across_sessions,
            )
        )
    return counts


def _count_rows_in_export(file_path: Path) -> int:
    """Count all data rows across all sheets in an export workbook."""
    from .excel_parser import list_sheets

    total = 0
    for sheet in list_sheets(file_path):
        total += len(parse_sheet(file_path, sheet))
    return total


def _default_backup_filename(config: dict) -> str:
    """Build a Beacon-style timestamped filename for full backups."""
    timestamp = datetime.now().strftime("%Y%m%d%H%M")
    site_name = str(config.get("beacon", {}).get("site_name", "Beacon")).strip()
    site_name = _sanitize_filename_fragment(site_name) or "Beacon"
    return f"{timestamp}_{site_name} u3abackup.xlsx"


def _sanitize_filename_fragment(value: str) -> str:
    """Remove characters that are invalid in Windows filenames."""
    invalid_chars = '<>:"/\\|?*'
    sanitized = "".join("_" if char in invalid_chars else char for char in value)
    return sanitized.strip().rstrip(".")


def _is_truthy(value: object) -> bool:
    """Parse common INI-style truthy values."""
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "on"}

