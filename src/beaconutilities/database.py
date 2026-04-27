"""SQLite staging storage for Beacon Excel workbooks.

Each sheet in each downloaded workbook is loaded into a separate SQLite table.
The table schema is derived at runtime from the Excel column headers — the
workbook structure drives the database structure with no configuration required.

Table naming
------------
Sheet names are sanitised to valid SQLite identifiers (lower-cased, runs of
non-alphanumeric characters collapsed to a single underscore).  Examples::

    Members          → members
    Group members    → group_members
    Group Ledgers    → group_ledgers

An automatically-managed column ``_staged_at`` (ISO timestamp, DEFAULT
CURRENT_TIMESTAMP) is prepended to every table so that rows from different sync
runs can be distinguished when running in persistent mode.

Persistence modes
-----------------
``append=False`` (default / session mode)
    Each sheet's table is dropped and recreated before inserting — only the
    current run's data is present after a sync.

``append=True`` (persistent mode)
    Tables are created if absent; new rows are inserted without clearing old
    ones.  Use ``_staged_at`` to filter by run.
"""

from __future__ import annotations

import re
import sqlite3
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def load_workbook_to_db(
    db_path: Path,
    workbook_path: Path,
    *,
    append: bool = False,
) -> dict[str, int]:
    """Load every sheet of an Excel workbook into SQLite tables.

    Args:
        db_path:       Path to the SQLite database file.  Parent directories
                       are created automatically.
        workbook_path: Path to the ``.xlsx`` file to load.
        append:        When ``False`` (default) each sheet's table is dropped
                       and recreated before inserting.  When ``True`` rows are
                       appended to existing tables (creating them if absent).

    Returns:
        A ``dict`` mapping sanitised table name → number of data rows inserted
        for that sheet.  Header rows are not counted.
    """
    db_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    counts: dict[str, int] = {}

    try:
        with sqlite3.connect(db_path) as conn:
            for sheet_name in wb.sheetnames:
                table = _sanitize(sheet_name)
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)

                try:
                    raw_headers = next(rows_iter)
                except StopIteration:
                    counts[table] = 0
                    continue

                columns = [
                    _sanitize(str(h)) if h is not None else f"col_{i}"
                    for i, h in enumerate(raw_headers)
                ]

                if not append:
                    conn.execute(f'DROP TABLE IF EXISTS "{table}"')

                col_defs = (
                    '"_staged_at" TEXT DEFAULT CURRENT_TIMESTAMP, '
                    + ", ".join(f'"{c}" TEXT' for c in columns)
                )
                conn.execute(
                    f'CREATE TABLE IF NOT EXISTS "{table}" ({col_defs})'
                )

                col_list = ", ".join(f'"{c}"' for c in columns)
                placeholders = ", ".join("?" * len(columns))
                row_count = 0
                for row in rows_iter:
                    values = [
                        str(v) if v is not None else None
                        for v in row[: len(columns)]
                    ]
                    conn.execute(
                        f'INSERT INTO "{table}" ({col_list}) VALUES ({placeholders})',
                        values,
                    )
                    row_count += 1

                counts[table] = row_count
    finally:
        wb.close()

    return counts


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _sanitize(name: str) -> str:
    """Return *name* as a safe, lower-cased SQLite identifier.

    Runs of characters that are not letters, digits, or underscores are
    collapsed into a single underscore; leading/trailing underscores are
    stripped.

    >>> _sanitize("Group members")
    'group_members'
    >>> _sanitize("e-mail")
    'e_mail'
    """
    name = str(name).strip().lower()
    name = re.sub(r"[^a-z0-9]+", "_", name)
    return name.strip("_") or "col"