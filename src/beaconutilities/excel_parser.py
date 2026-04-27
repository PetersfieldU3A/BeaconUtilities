"""Excel file parsing for BeaconUtilities.

Reads Beacon-exported .xlsx files into plain Python dicts, one per data row,
with the first row of each sheet treated as column headers.  No assumptions
are made about field names — the caller decides which sheets and columns are
meaningful.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl


def list_sheets(file_path: Path) -> list[str]:
    """Return all sheet names in an Excel workbook.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        Ordered list of sheet names as they appear in the workbook.

    Raises:
        FileNotFoundError: If *file_path* does not exist.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    wb = openpyxl.load_workbook(file_path, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def parse_sheet(file_path: Path, sheet_name: str | int = 0) -> list[dict]:
    """Parse one sheet from an Excel workbook into a list of row dicts.

    The first row is treated as column headers.  Completely empty rows are
    skipped.  Cell values are returned as-is (str, int, float, datetime,
    or None).

    Args:
        file_path: Path to the .xlsx file.
        sheet_name: Sheet name (str) or zero-based index (int).
                    Defaults to the first sheet.

    Returns:
        List of dicts mapping header names to cell values.

    Raises:
        FileNotFoundError: If *file_path* does not exist.
        KeyError: If a named sheet is not found in the workbook.
        IndexError: If a numeric index is out of range.
        ValueError: If the sheet contains no header row.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        if isinstance(sheet_name, int):
            if sheet_name >= len(wb.worksheets):
                raise IndexError(
                    f"Sheet index {sheet_name} out of range "
                    f"(workbook has {len(wb.worksheets)} sheets)"
                )
            ws = wb.worksheets[sheet_name]
        else:
            if sheet_name not in wb.sheetnames:
                raise KeyError(
                    f"Sheet '{sheet_name}' not found in {file_path}. "
                    f"Available sheets: {wb.sheetnames}"
                )
            ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return []

    header_row = rows[0]
    if all(h is None for h in header_row):
        raise ValueError(f"Sheet '{sheet_name}' in {file_path} has no header row")

    # Build header list; unnamed columns get a generated name to avoid key collisions
    headers: list[str] = []
    for i, h in enumerate(header_row):
        headers.append(str(h).strip() if h is not None else f"_col_{i}")

    result: list[dict] = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        # Pad short rows so all header keys are present
        padded = list(row) + [None] * (len(headers) - len(row))
        result.append({headers[i]: padded[i] for i in range(len(headers))})

    return result
