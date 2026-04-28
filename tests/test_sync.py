"""Tests for beaconutilities.sync — orchestration layer.

All external I/O (Playwright, WordPress HTTP) is mocked so tests run
offline and without side effects.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from beaconutilities.sync import (
    run_beacon_full_backup,
    run_beacon_to_sqlite_dry_run,
    run_export_member_names,
    run_sync,
)


def _write_excel(path: Path, rows: list[list], sheet_name: str = "Sheet") -> None:
    """Helper: write a single-sheet .xlsx at *path*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


class TestRunSyncPreflight:
    def test_returns_preflight_failed_on_bad_beacon_config(self, tmp_path):
        cfg = {"beacon": {}, "beacon_export": {}, "wordpress": {}}
        result = run_sync(cfg)
        assert result["status"] == "preflight_failed"

    def test_returns_preflight_failed_on_bad_wordpress_config(self, minimal_config):
        minimal_config["wordpress"]["site_url"] = ""
        with patch("beaconutilities.sync.download_beacon_exports") as mock_dl:
            # preflight_wordpress should fail before download is called
            result = run_sync(minimal_config, dry_run=False)
        assert result["status"] == "preflight_failed"
        mock_dl.assert_not_called()


class TestRunSyncDownloadFailure:
    def test_returns_download_failed_on_exception(self, minimal_config):
        with patch(
            "beaconutilities.sync.download_beacon_exports",
            side_effect=RuntimeError("Connection refused"),
        ):
            result = run_sync(minimal_config)
        assert result["status"] == "download_failed"
        assert "Connection refused" in result["errors"][0]


class TestRunSyncDryRun:
    def test_dry_run_skips_wordpress(self, minimal_config, tmp_path):
        members_xlsx = tmp_path / "members.xlsx"
        groups_xlsx = tmp_path / "groups.xlsx"
        _write_excel(members_xlsx, [
            ["Member No", "First Name", "Last Name", "Email"],
            ["1001", "Alice", "Smith", "alice@example.com"],
        ])
        _write_excel(groups_xlsx, [
            ["Group Id", "Group Name", "Description"],
            ["G1", "Photography", "Camera fans"],
        ])

        with patch(
            "beaconutilities.sync.download_beacon_exports",
            return_value={"members": members_xlsx, "groups": groups_xlsx},
        ):
            with patch("beaconutilities.sync.client_from_config") as mock_wp:
                result = run_sync(minimal_config, dry_run=True)

        mock_wp.assert_not_called()
        assert result["dry_run"] is True
        assert result["members_extracted"] == 1
        assert result["groups_extracted"] == 1
        assert result["staged"] == 0
        assert result["published"] == 0

    def test_dry_run_does_not_update_state(self, minimal_config, tmp_path):
        members_xlsx = tmp_path / "members.xlsx"
        groups_xlsx = tmp_path / "groups.xlsx"
        _write_excel(members_xlsx, [["Member No"], ["1001"]])
        _write_excel(groups_xlsx, [["Group Id"], ["G1"]])

        with patch(
            "beaconutilities.sync.download_beacon_exports",
            return_value={"members": members_xlsx, "groups": groups_xlsx},
        ):
            with patch("beaconutilities.sync.save_state") as mock_save:
                run_sync(minimal_config, dry_run=True)

        mock_save.assert_not_called()


class TestRunSyncLivePublish:
    def test_publishes_all_records(self, minimal_config, tmp_path):
        members_xlsx = tmp_path / "members.xlsx"
        groups_xlsx = tmp_path / "groups.xlsx"
        _write_excel(members_xlsx, [
            ["Member No", "First Name", "Last Name", "Email"],
            ["1001", "Alice", "Smith", "alice@example.com"],
            ["1002", "Bob", "Jones", "bob@example.com"],
        ])
        _write_excel(groups_xlsx, [
            ["Group Id", "Group Name", "Description"],
            ["G1", "Photography", ""],
        ])

        mock_wp_client = MagicMock()
        mock_wp_client.upsert_post.return_value = {"id": 1}

        with patch(
            "beaconutilities.sync.download_beacon_exports",
            return_value={"members": members_xlsx, "groups": groups_xlsx},
        ):
            with patch("beaconutilities.sync.client_from_config", return_value=mock_wp_client):
                with patch("beaconutilities.sync.save_state"):
                    result = run_sync(minimal_config, dry_run=False)

        assert result["status"] == "ok"
        assert result["published"] == 3  # 2 members + 1 group
        assert mock_wp_client.upsert_post.call_count == 3
        assert result["staged"] == 0

    def test_partial_status_on_publish_error(self, minimal_config, tmp_path):
        members_xlsx = tmp_path / "members.xlsx"
        groups_xlsx = tmp_path / "groups.xlsx"
        _write_excel(members_xlsx, [["Member No"], ["1001"]])
        _write_excel(groups_xlsx, [["Group Id"], ["G1"]])

        mock_wp_client = MagicMock()
        mock_wp_client.upsert_post.side_effect = Exception("HTTP 500")

        with patch(
            "beaconutilities.sync.download_beacon_exports",
            return_value={"members": members_xlsx, "groups": groups_xlsx},
        ):
            with patch("beaconutilities.sync.client_from_config", return_value=mock_wp_client):
                with patch("beaconutilities.sync.save_state"):
                    result = run_sync(minimal_config, dry_run=False)

        assert result["status"] == "partial"
        assert len(result["errors"]) > 0


class TestRunSyncDatabaseStaging:
    def test_stages_records_when_enabled(self, minimal_config, tmp_path):
        minimal_config["database"] = {
            "enabled": "true",
            "path": str(tmp_path / "state" / "beacon_data.db"),
        }
        members_xlsx = tmp_path / "members.xlsx"
        groups_xlsx = tmp_path / "groups.xlsx"
        _write_excel(members_xlsx, [["Member No"], ["1001"]], sheet_name="members")
        _write_excel(groups_xlsx, [["Group Id"], ["G1"]], sheet_name="groups")

        mock_wp_client = MagicMock()
        mock_wp_client.upsert_post.return_value = {"id": 1}

        with patch(
            "beaconutilities.sync.download_beacon_exports",
            return_value={"members": members_xlsx, "groups": groups_xlsx},
        ):
            with patch("beaconutilities.sync.client_from_config", return_value=mock_wp_client):
                with patch("beaconutilities.sync.save_state"):
                    result = run_sync(minimal_config, dry_run=False)

        assert result["staged"] == 2

    def test_non_persistent_mode_replaces_rows_each_run(self, minimal_config, tmp_path):
        db_path = tmp_path / "state" / "beacon_data.db"
        minimal_config["database"] = {
            "enabled": "true",
            "path": str(db_path),
            "persist_across_sessions": "false",
        }
        members_xlsx = tmp_path / "members.xlsx"
        groups_xlsx = tmp_path / "groups.xlsx"
        _write_excel(members_xlsx, [["Member No"], ["1001"]], sheet_name="members")
        _write_excel(groups_xlsx, [["Group Id"], ["G1"]], sheet_name="groups")

        with patch(
            "beaconutilities.sync.download_beacon_exports",
            return_value={"members": members_xlsx, "groups": groups_xlsx},
        ):
            first = run_sync(minimal_config, dry_run=True)
        with patch(
            "beaconutilities.sync.download_beacon_exports",
            return_value={"members": members_xlsx, "groups": groups_xlsx},
        ):
            second = run_sync(minimal_config, dry_run=True)

        assert first["staged"] == 2
        assert second["staged"] == 2

        import sqlite3

        # Session mode: DROP+recreate each run — only 1 row per table, 2 total
        with sqlite3.connect(db_path) as conn:
            mem_count = conn.execute("SELECT COUNT(*) FROM members").fetchone()[0]
            grp_count = conn.execute("SELECT COUNT(*) FROM groups").fetchone()[0]
        assert mem_count + grp_count == 2

    def test_persistent_mode_accumulates_rows(self, minimal_config, tmp_path):
        db_path = tmp_path / "state" / "beacon_data.db"
        minimal_config["database"] = {
            "enabled": "true",
            "path": str(db_path),
            "persist_across_sessions": "true",
        }
        members_xlsx = tmp_path / "members.xlsx"
        groups_xlsx = tmp_path / "groups.xlsx"
        _write_excel(members_xlsx, [["Member No"], ["1001"]], sheet_name="members")
        _write_excel(groups_xlsx, [["Group Id"], ["G1"]], sheet_name="groups")

        with patch(
            "beaconutilities.sync.download_beacon_exports",
            return_value={"members": members_xlsx, "groups": groups_xlsx},
        ):
            run_sync(minimal_config, dry_run=True)
        with patch(
            "beaconutilities.sync.download_beacon_exports",
            return_value={"members": members_xlsx, "groups": groups_xlsx},
        ):
            run_sync(minimal_config, dry_run=True)

        import sqlite3

        # Persistent mode: append=True — 2 rows per table × 2 tables = 4 total
        with sqlite3.connect(db_path) as conn:
            mem_count = conn.execute("SELECT COUNT(*) FROM members").fetchone()[0]
            grp_count = conn.execute("SELECT COUNT(*) FROM groups").fetchone()[0]
        assert mem_count + grp_count == 4

    def test_sets_partial_if_staging_fails(self, minimal_config, tmp_path):
        minimal_config["database"] = {
            "enabled": "true",
            "path": str(tmp_path / "state" / "beacon_data.db"),
        }
        members_xlsx = tmp_path / "members.xlsx"
        groups_xlsx = tmp_path / "groups.xlsx"
        _write_excel(members_xlsx, [["Member No"], ["1001"]])
        _write_excel(groups_xlsx, [["Group Id"], ["G1"]])

        mock_wp_client = MagicMock()
        mock_wp_client.upsert_post.return_value = {"id": 1}

        with patch(
            "beaconutilities.sync.download_beacon_exports",
            return_value={"members": members_xlsx, "groups": groups_xlsx},
        ):
            with patch("beaconutilities.sync.client_from_config", return_value=mock_wp_client):
                with patch("beaconutilities.sync.load_workbook_to_db", side_effect=RuntimeError("db down")):
                    with patch("beaconutilities.sync.save_state"):
                        result = run_sync(minimal_config, dry_run=False)

        assert result["status"] == "partial"
        assert any("db_staging" in msg for msg in result["errors"])

    def test_state_saved_after_live_sync(self, minimal_config, tmp_path):
        members_xlsx = tmp_path / "members.xlsx"
        groups_xlsx = tmp_path / "groups.xlsx"
        _write_excel(members_xlsx, [["Member No"], ["1001"]])
        _write_excel(groups_xlsx, [["Group Id"], ["G1"]])

        mock_wp_client = MagicMock()
        mock_wp_client.upsert_post.return_value = {"id": 1}

        with patch(
            "beaconutilities.sync.download_beacon_exports",
            return_value={"members": members_xlsx, "groups": groups_xlsx},
        ):
            with patch("beaconutilities.sync.client_from_config", return_value=mock_wp_client):
                with patch("beaconutilities.sync.save_state") as mock_save:
                    with patch("beaconutilities.sync.load_state", return_value={}):
                        run_sync(minimal_config, dry_run=False)

        mock_save.assert_called_once()
        saved_state = mock_save.call_args[0][1]
        assert "last_sync" in saved_state


class TestRunBeaconToSqliteDryRun:
    def test_stages_downloaded_workbooks(self, minimal_config, tmp_path):
        db_path = tmp_path / "state" / "beacon_data.db"
        minimal_config["database"] = {
            "path": str(db_path),
            "persist_across_sessions": "false",
        }

        members_xlsx = tmp_path / "members.xlsx"
        groups_xlsx = tmp_path / "groups.xlsx"
        _write_excel(members_xlsx, [["mem_no", "forename"], ["1001", "Alice"]], sheet_name="Members")
        _write_excel(groups_xlsx, [["gkey", "group_name"], ["G1", "Photography"]], sheet_name="Groups")

        with patch(
            "beaconutilities.sync.download_beacon_exports",
            return_value={"members": members_xlsx, "groups": groups_xlsx},
        ):
            result = run_beacon_to_sqlite_dry_run(minimal_config)

        assert result["status"] == "ok"
        assert result["members_extracted"] == 1
        assert result["groups_extracted"] == 1
        assert result["staged"] == 2
        assert result["tables"] == 2


class TestRunExportMemberNames:
    def test_writes_member_names_workbook_with_required_columns(self, minimal_config, tmp_path):
        members_xlsx = tmp_path / "members.xlsx"
        groups_xlsx = tmp_path / "groups.xlsx"
        _write_excel(
            members_xlsx,
            [
                ["mem_no", "status", "title", "forename", "surname", "e-mail"],
                ["1001", "Current", "Mr", "Alan", "Brown", "a@example.com"],
                ["1002", "Current", "Ms", "Jane", "Smith", "j@example.com"],
            ],
            sheet_name="Members",
        )
        _write_excel(groups_xlsx, [["gkey"], ["G1"]], sheet_name="Groups")

        output_dir = tmp_path / "exports"
        with patch(
            "beaconutilities.sync.download_beacon_exports",
            return_value={"members": members_xlsx, "groups": groups_xlsx},
        ):
            result = run_export_member_names(minimal_config, output_dir=output_dir)

        assert result["status"] == "ok"
        assert result["rows_written"] == 2

        out_file = output_dir / "Member_Names.xlsx"
        assert out_file.exists()

        wb = openpyxl.load_workbook(out_file, read_only=True)
        assert wb.sheetnames == ["Member Names"]
        ws = wb["Member Names"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        assert rows[0] == ("mem_no", "status", "title", "forename", "surname")
        assert rows[1] == ("1001", "Current", "Mr", "Alan", "Brown")
        assert rows[2] == ("1002", "Current", "Ms", "Jane", "Smith")

    def test_fails_when_required_member_columns_are_missing(self, minimal_config, tmp_path):
        members_xlsx = tmp_path / "members.xlsx"
        groups_xlsx = tmp_path / "groups.xlsx"
        _write_excel(
            members_xlsx,
            [["mem_no", "forename", "surname"], ["1001", "Alan", "Brown"]],
            sheet_name="Members",
        )
        _write_excel(groups_xlsx, [["gkey"], ["G1"]], sheet_name="Groups")

        with patch(
            "beaconutilities.sync.download_beacon_exports",
            return_value={"members": members_xlsx, "groups": groups_xlsx},
        ):
            result = run_export_member_names(minimal_config, output_dir=tmp_path)

        assert result["status"] == "parse_failed"
        assert any("Missing required member columns" in msg for msg in result["errors"])


class TestRunBeaconFullBackup:
    def test_downloads_backup_to_configured_directory(self, minimal_config, tmp_path):
        expected = tmp_path / "outputs" / "202604281316_TestU3A u3abackup.xlsx"
        with patch("beaconutilities.sync.datetime") as mock_datetime:
            mock_datetime.now.return_value = datetime(2026, 4, 28, 13, 16)
            with patch(
                "beaconutilities.sync.download_beacon_backup",
                return_value=expected,
            ) as mock_download:
                result = run_beacon_full_backup(minimal_config)

        assert result["status"] == "ok"
        assert result["output_file"] == str(expected)
        mock_download.assert_called_once()
        called_destination = mock_download.call_args[0][1]
        assert called_destination == expected

    def test_output_file_override_accepts_directory(self, minimal_config, tmp_path):
        override_dir = tmp_path / "alt-backups"
        expected = override_dir / "202604281316_TestU3A u3abackup.xlsx"
        with patch("beaconutilities.sync.datetime") as mock_datetime:
            mock_datetime.now.return_value = datetime(2026, 4, 28, 13, 16)
            with patch(
                "beaconutilities.sync.download_beacon_backup",
                return_value=expected,
            ) as mock_download:
                result = run_beacon_full_backup(minimal_config, output_file=override_dir)

        assert result["status"] == "ok"
        assert result["output_file"] == str(expected)
        called_destination = mock_download.call_args[0][1]
        assert called_destination == expected

    def test_output_file_override_accepts_file_path(self, minimal_config, tmp_path):
        output_file = tmp_path / "backups" / "manual_name.xlsx"
        with patch(
            "beaconutilities.sync.download_beacon_backup",
            return_value=output_file,
        ) as mock_download:
            result = run_beacon_full_backup(minimal_config, output_file=output_file)

        assert result["status"] == "ok"
        assert result["output_file"] == str(output_file)
        called_destination = mock_download.call_args[0][1]
        assert called_destination == output_file

    def test_default_backup_filename_sanitizes_invalid_site_name(self, minimal_config, tmp_path):
        minimal_config["beacon"]["site_name"] = 'Test/U3A: North'
        expected = tmp_path / "outputs" / "202604281316_Test_U3A_ North u3abackup.xlsx"
        with patch("beaconutilities.sync.datetime") as mock_datetime:
            mock_datetime.now.return_value = datetime(2026, 4, 28, 13, 16)
            with patch(
                "beaconutilities.sync.download_beacon_backup",
                return_value=expected,
            ) as mock_download:
                result = run_beacon_full_backup(minimal_config)

        assert result["status"] == "ok"
        called_destination = mock_download.call_args[0][1]
        assert called_destination == expected

    def test_returns_download_failed_on_backup_exception(self, minimal_config):
        with patch(
            "beaconutilities.sync.download_beacon_backup",
            side_effect=RuntimeError("Backup link not found"),
        ):
            result = run_beacon_full_backup(minimal_config)

        assert result["status"] == "download_failed"
        assert "Backup link not found" in result["errors"][0]
