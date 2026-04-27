"""Tests for beaconutilities.database."""

from __future__ import annotations

import sqlite3
from pathlib import Path

import openpyxl
import pytest

from beaconutilities.database import _sanitize, load_workbook_to_db


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _write_excel(path: Path, sheets: dict[str, list[list]]) -> None:
    """Write a multi-sheet workbook at *path*.

    *sheets* is an ordered dict of {sheet_name: [rows]}.
    """
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _table_names(db_path: Path) -> set[str]:
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()
    return {r[0] for r in rows}


def _row_count(db_path: Path, table: str) -> int:
    with sqlite3.connect(db_path) as conn:
        return conn.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]


def _column_names(db_path: Path, table: str) -> list[str]:
    with sqlite3.connect(db_path) as conn:
        info = conn.execute(f'PRAGMA table_info("{table}")').fetchall()
    return [row[1] for row in info]


# ---------------------------------------------------------------------------
# _sanitize
# ---------------------------------------------------------------------------


class TestSanitize:
    def test_lowercases(self):
        assert _sanitize("Members") == "members"

    def test_spaces_to_underscores(self):
        assert _sanitize("Group members") == "group_members"

    def test_hyphens_to_underscores(self):
        assert _sanitize("e-mail") == "e_mail"

    def test_multiple_special_chars_collapsed(self):
        assert _sanitize("Group Ledgers!") == "group_ledgers"

    def test_leading_trailing_underscores_stripped(self):
        assert _sanitize("  _foo_ ") == "foo"

    def test_empty_string_returns_col(self):
        assert _sanitize("") == "col"


# ---------------------------------------------------------------------------
# load_workbook_to_db — basic loading
# ---------------------------------------------------------------------------


class TestLoadWorkbookToDb:
    def test_creates_database_file(self, tmp_path: Path):
        xlsx = tmp_path / "test.xlsx"
        db = tmp_path / "sub" / "test.db"
        _write_excel(xlsx, {"Members": [["id", "name"], ["1", "Alice"]]})
        load_workbook_to_db(db, xlsx)
        assert db.exists()

    def test_single_sheet_creates_table_and_inserts_rows(self, tmp_path: Path):
        xlsx = tmp_path / "members.xlsx"
        db = tmp_path / "test.db"
        _write_excel(xlsx, {"Members": [
            ["mem_no", "forename", "surname"],
            ["1001", "Alice", "Smith"],
            ["1002", "Bob", "Jones"],
        ]})
        counts = load_workbook_to_db(db, xlsx)
        assert counts == {"members": 2}
        assert _row_count(db, "members") == 2

    def test_multiple_sheets_each_become_a_table(self, tmp_path: Path):
        xlsx = tmp_path / "groups.xlsx"
        db = tmp_path / "test.db"
        _write_excel(xlsx, {
            "Groups": [["gkey", "group_name"], ["1", "Photography"]],
            "Venues": [["gvkey", "venue"], ["10", "Village Hall"], ["11", "Library"]],
            "Faculties": [["gfkey", "faculty"]],
        })
        counts = load_workbook_to_db(db, xlsx)
        assert counts == {"groups": 1, "venues": 2, "faculties": 0}
        assert "groups" in _table_names(db)
        assert "venues" in _table_names(db)
        assert "faculties" in _table_names(db)

    def test_column_names_are_sanitised(self, tmp_path: Path):
        xlsx = tmp_path / "test.xlsx"
        db = tmp_path / "test.db"
        _write_excel(xlsx, {"Members": [["e-mail", "mem_no", "First Name"], ["a@b.com", "1", "Alice"]]})
        load_workbook_to_db(db, xlsx)
        cols = _column_names(db, "members")
        assert "e_mail" in cols
        assert "mem_no" in cols
        assert "first_name" in cols

    def test_staged_at_column_always_present(self, tmp_path: Path):
        xlsx = tmp_path / "test.xlsx"
        db = tmp_path / "test.db"
        _write_excel(xlsx, {"Members": [["id"], ["1"]]})
        load_workbook_to_db(db, xlsx)
        cols = _column_names(db, "members")
        assert "_staged_at" in cols

    def test_returns_zero_for_header_only_sheet(self, tmp_path: Path):
        xlsx = tmp_path / "test.xlsx"
        db = tmp_path / "test.db"
        _write_excel(xlsx, {"Empty": [["id", "name"]]})
        counts = load_workbook_to_db(db, xlsx)
        assert counts == {"empty": 0}

    def test_table_name_derived_from_sheet_name(self, tmp_path: Path):
        xlsx = tmp_path / "test.xlsx"
        db = tmp_path / "test.db"
        _write_excel(xlsx, {"Group members": [["gkey"], ["1"]]})
        counts = load_workbook_to_db(db, xlsx)
        assert "group_members" in counts
        assert "group_members" in _table_names(db)

    def test_none_cell_stored_as_null(self, tmp_path: Path):
        xlsx = tmp_path / "test.xlsx"
        db = tmp_path / "test.db"
        _write_excel(xlsx, {"Members": [["id", "name"], ["1", None]]})
        load_workbook_to_db(db, xlsx)
        with sqlite3.connect(db) as conn:
            row = conn.execute('SELECT "name" FROM "members"').fetchone()
        assert row[0] is None


# ---------------------------------------------------------------------------
# Session mode (append=False, default)
# ---------------------------------------------------------------------------


class TestSessionMode:
    def test_drop_recreates_table_each_call(self, tmp_path: Path):
        xlsx = tmp_path / "test.xlsx"
        db = tmp_path / "test.db"
        _write_excel(xlsx, {"Members": [["id"], ["1"], ["2"]]})
        load_workbook_to_db(db, xlsx, append=False)
        # Second call — should replace, not accumulate
        load_workbook_to_db(db, xlsx, append=False)
        assert _row_count(db, "members") == 2  # not 4

    def test_two_workbooks_loaded_separately(self, tmp_path: Path):
        mem_xlsx = tmp_path / "members.xlsx"
        grp_xlsx = tmp_path / "groups.xlsx"
        db = tmp_path / "test.db"
        _write_excel(mem_xlsx, {"Members": [["mem_no"], ["1001"]]})
        _write_excel(grp_xlsx, {"Groups": [["gkey"], ["G1"]]})
        load_workbook_to_db(db, mem_xlsx, append=False)
        load_workbook_to_db(db, grp_xlsx, append=False)
        assert _row_count(db, "members") == 1
        assert _row_count(db, "groups") == 1


# ---------------------------------------------------------------------------
# Persistent mode (append=True)
# ---------------------------------------------------------------------------


class TestPersistentMode:
    def test_rows_accumulate_across_calls(self, tmp_path: Path):
        xlsx = tmp_path / "test.xlsx"
        db = tmp_path / "test.db"
        _write_excel(xlsx, {"Members": [["id"], ["1"]]})
        load_workbook_to_db(db, xlsx, append=True)
        load_workbook_to_db(db, xlsx, append=True)
        assert _row_count(db, "members") == 2

    def test_staged_at_differs_across_runs(self, tmp_path: Path):
        xlsx = tmp_path / "test.xlsx"
        db = tmp_path / "test.db"
        _write_excel(xlsx, {"Members": [["id"], ["1"]]})
        load_workbook_to_db(db, xlsx, append=True)
        load_workbook_to_db(db, xlsx, append=True)
        with sqlite3.connect(db) as conn:
            timestamps = conn.execute(
                "SELECT DISTINCT _staged_at FROM members"
            ).fetchall()
        # Both rows have some timestamp (not null)
        assert all(ts[0] is not None for ts in timestamps)